"""
导出服务 - 导出 Excel 文件
"""
import io
from datetime import datetime
from typing import List, Optional
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.models import User, Work, Comment


class ExporterService:
    """导出服务"""

    def __init__(self):
        # 样式定义
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, size=11, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _set_column_widths(self, ws, widths: List[int]):
        """设置列宽"""
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _style_header(self, ws, row: int, col_count: int):
        """设置表头样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.alignment
            cell.border = self.border

    def _style_data_row(self, ws, row: int, col_count: int):
        """设置数据行样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = self.alignment
            cell.border = self.border

    def export_works(self, db: Session, work_ids: List[str] = None,
                     limit: int = 1000) -> io.BytesIO:
        """导出作品到 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "作品数据"

        headers = [
            "作品ID", "作品URL", "类型", "标题", "描述",
            "点赞数", "评论数", "收藏数", "分享数", "赞赏数",
            "视频链接", "图片链接", "话题",
            "作者昵称", "作者主页", "发布时间", "采集时间"
        ]
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        self._set_column_widths(ws, [20, 40, 8, 40, 40, 10, 10, 10, 10, 10, 50, 50, 30, 15, 40, 20, 20])

        query = db.query(Work).join(User, isouter=True)
        if work_ids:
            query = query.filter(Work.work_id.in_(work_ids))
        works = query.order_by(Work.crawled_at.desc()).limit(limit).all()

        for work in works:
            images_str = ""
            if work.images:
                images_str = "\n".join(work.images) if isinstance(work.images, list) else str(work.images)
            topics_str = ""
            if work.topics:
                topics_str = ", ".join(work.topics) if isinstance(work.topics, list) else str(work.topics)

            ws.append([
                work.work_id,
                work.work_url or "",
                "视频" if work.work_type == "video" else "图集",
                work.title or "",
                work.description or "",
                work.digg_count,
                work.comment_count,
                work.collect_count,
                work.share_count,
                work.admire_count,
                work.video_url or "",
                images_str,
                topics_str,
                work.user.nickname if work.user else "",
                work.user.user_url if work.user else "",
                work.create_time.strftime("%Y-%m-%d %H:%M:%S") if work.create_time else "",
                work.crawled_at.strftime("%Y-%m-%d %H:%M:%S") if work.crawled_at else "",
            ])
            self._style_data_row(ws, ws.max_row, len(headers))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_users(self, db: Session, uids: List[str] = None,
                     limit: int = 1000) -> io.BytesIO:
        """导出用户到 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "用户数据"

        headers = [
            "用户ID", "抖音号", "昵称", "性别", "年龄", "IP属地",
            "粉丝数", "关注数", "作品数", "获赞数",
            "简介", "主页URL", "更新时间"
        ]
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        self._set_column_widths(ws, [20, 15, 20, 8, 8, 12, 12, 12, 10, 12, 40, 45, 20])

        query = db.query(User)
        if uids:
            query = query.filter(User.uid.in_(uids))
        users = query.order_by(User.updated_at.desc()).limit(limit).all()

        for user in users:
            gender = "未知"
            if user.gender == 1:
                gender = "男"
            elif user.gender == 2:
                gender = "女"

            ws.append([
                user.uid,
                user.unique_id or "",
                user.nickname or "",
                gender,
                user.age or "",
                user.ip_location or "",
                user.follower_count,
                user.following_count,
                user.aweme_count,
                user.total_favorited or user.favorited_count,
                user.signature or "",
                user.user_url or "",
                user.updated_at.strftime("%Y-%m-%d %H:%M:%S") if user.updated_at else "",
            ])
            self._style_data_row(ws, ws.max_row, len(headers))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_comments(self, db: Session, work_id: str = None,
                        limit: int = 1000) -> io.BytesIO:
        """导出评论到 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "评论数据"

        # 表头
        headers = [
            "评论ID", "作品ID", "用户昵称", "评论内容",
            "点赞数", "评论时间", "采集时间"
        ]
        ws.append(headers)
        self._style_header(ws, 1, len(headers))
        self._set_column_widths(ws, [25, 25, 20, 60, 12, 20, 20])

        # 查询数据
        query = db.query(Comment).join(Work).join(User, isouter=True)
        if work_id:
            query = query.filter(Work.work_id == work_id)
        comments = query.order_by(Comment.crawled_at.desc()).limit(limit).all()

        # 写入数据
        for comment in comments:
            ws.append([
                comment.comment_id,
                comment.work.work_id if comment.work else "",
                comment.user.nickname if comment.user else "",
                comment.content or "",
                comment.digg_count,
                comment.create_time.strftime("%Y-%m-%d %H:%M:%S") if comment.create_time else "",
                comment.crawled_at.strftime("%Y-%m-%d %H:%M:%S") if comment.crawled_at else "",
            ])
            self._style_data_row(ws, ws.max_row, len(headers))

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# 全局导出服务实例
exporter = ExporterService()


def get_exporter() -> ExporterService:
    """获取导出服务实例"""
    return exporter
